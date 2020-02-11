''' Asha Check '''

import time
import getpass
import requests
import openpyxl
from pathlib import Path
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

'''''''''''''''''''''
#User Input
'''''''''''''''''''''
pcnam = getpass.getuser()
print('\nWelcome '+pcnam+'\n')

#print('\nEnter SAP ID: ')
#sapid='3228'#input()

'''''''''''''''''''''
#Chrome Driver Loads
'''''''''''''''''''''
chrdrv = "C:\\Users\\{0}\\Downloads\\chromedriver.exe".format(pcnam)
url = "http://www.mumbainews24x7.com/Abhishek/chromedriver.exe"
if Path(chrdrv).is_file():
    print()
else:
    r = requests.get(url, stream = True)
    with open(file, 'wb') as f:
        for chunk in r.iter_content(chunk_size=1024):
            f.write(chunk)

'''''''''''
#Py Drivers
'''''''''''

opts = Options()
opts.add_argument('start-maximized')
opts.add_argument('disable-infobars')
opts.add_argument('--ignore-certificate-errors')
opts.add_argument("--test-type")

prefs = {'download.default_directory' : 'C:\\Users\\Abhishek.Chitnis\\Desktop\\Auto'}
opts.add_experimental_option('prefs', prefs)
#driver = webdriver.Chrome(options=opts, executable_path=chrdrv)

path="C:\\Users\\Abhishek.Chitnis\Desktop\\Auto\\ACo.xlsx"

''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
# 16

#unit test frame work
import unittest
from selenium import webdriver
class searchengineertest(unittest.TestCase):

    def test_google(self):
        self.driver=webdriver.Chrome("C://Users//asha.colaco//RQM_Utility//RQM_Utility//Library//chromedriver.exe")
        self.driver.get("https://www.google.com/")
        print(self.driver.title)
        self.driver.close()


    def test_peace(self):
        self.driver=webdriver.Chrome("C://Users//asha.colaco//RQM_Utility//RQM_Utility//Library//chromedriver.exe")
        self.driver.get("https://www.ajio.com/")
        print(self.driver.title)
        self.driver.close()

if __name__ =="__main__":
    unittest.main(argv=[''], verbosity=2, exit=False)

'''
#15
'''
def getrowcount():
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.active
    rows=sheet.max_row
    #print(rows)
    return rows

def getcolumncount():
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.active
    cols=sheet.max_column
    #print(cols)
    return cols
# you need to store the value in a variable when u use return    
rows=getrowcount()
cols=getcolumncount()
print(rows,cols)

#writting the data in the excel
def writting():
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.active
    for r in range(1,6):
        for c in range(1,5):
            sheet.cell(4,4).value="hello"
    workbook.save(path)
    print(sheet.cell(4,4).value)
    
writting()

#14
'''
#load the workbook
workbook=openpyxl.load_workbook(path)

def getrowcount():
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.active
    rows=sheet.max_row
    print(rows)

def getcolumncount():
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.active
    cols=sheet.max_column
    print(cols) 

'''
#13
'''
def getrowcount():
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.active
    rows=sheet.max_row
    #return(sheet.max_row)
    print(rows)

def getcolumncount():
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.active
    cols=sheet.max_column
    print(cols)


getrowcount()
getcolumncount()

'''
#12
'''
#download in a particular location
from selenium import webdriver
import time
import openpyxl

path="C:\\Users\\asha.colaco\\Desktop\\colaco.xlsx"

#load the workbook
workbook=openpyxl.load_workbook(path)

#active the workbook
#sheet=workbook.get_sheet_by_name("Sheet2")
sheet=workbook.active
#row=r
#column=c

for r in range(1,6):
    for c in range(1,4):
        #print sheet.cell(row=r,column=c).value="welcome")
        #print(sheet.cell(row=r,column=c).value="welcome")
        sheet.cell(1,1).value="welcome"
        sheet.cell(1,2).value="we"
        
        
        
workbook.save(path)
'''
#11
'''
path="C:\\Users\\asha.colaco\\Desktop\\colaco.xlsx"

#load the workbook
workbook=openpyxl.load_workbook(path)

#active the workbook
#sheet=workbook.get_sheet_by_name("Sheet2")
sheet=workbook.active
row=r
column=c

for r in range(1,6):
    for c in range(1,4):
        #print sheet.cell(row=r,column=c).value="welcome")
        #print(sheet.cell(row=r,column=c).value="welcome")
        sheet.cell(row=r,column=c).value="welcome"

workbook.save(path)

'''
#10
'''
driver.get("http://demo.automationtesting.in/WebTable.html")
rows=driver.find_elements_by_xpath("//div[@class='ui-grid-row ng-scope']")
m=len(rows)
print(m)
print(rows)
cols=driver.find_elements_by_xpath("//div[@class='ui-grid-cell-contents ng-binding ng-scope']")
n=len(cols)
print(n)
print(cols)

#a=1,rows+1
#b=1,cols+1
for r in range(m):
    for c in range(n):
        print(m[r],n[c])
'''
#9
'''
prefs = {'download.default_directory' : 'C:\\Users\\Abhishek.Chitnis\\Desktop'}
opts.add_experimental_option('prefs', prefs)

driver = webdriver.Chrome(options=opts, executable_path=chrdrv)

driver.get("http://demo.automationtesting.in/FileDownload.html")
driver.find_element_by_id("textbox").send_keys("anjali")
driver.find_element_by_id("createTxt").click()
driver.find_element_by_id("link-to-download").click()
'''
#8
'''
from selenium import webdriver
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
driver=webdriver.Chrome("C://Users//Abhishek.Chitnis//Downloads//chromedriver.exe")

driver.get("https://the-internet.herokuapp.com/")
driver.maximize_window()
driver.find_element_by_xpath("//a[@href='/nested_frames']").click()
driver.switch_to_frame("frame-bottom")
new=driver.find_element_by_xpath("//*[contains(text(),'BOTTOM')]")
print(new.text)
'''

#7
'''
driver.get("https://the-internet.herokuapp.com/")
driver.maximize_window()
driver.find_element_by_xpath("//a[@href='/nested_frames']").click()
driver.switch_to_frame("frame-top")
driver.switch_to_frame("frame-middle")
asha=driver.find_element_by_xpath("//div[@id='content']")
print(asha.text)
'''
#6
'''
asha=[5,3]
for x in asha:
    if asha[x]==5 and 3:
        print("new")
    else:
        print("happy")

'''
#5th
'''
driver.get("https://rahulshettyacademy.com/seleniumPractise/#/")
driver.maximize_window()
asha=driver.find_elements_by_css_selector("h4.product-name")
print(len(asha))
for dil in asha:
       print(dil.text)
if dil.text=="Tomato - 1 Kg" or "Cucumber - 1 Kg":
       driver.find_element_by_xpath("//*[text()='Tomato - 1 Kg']/following-sibling::div[2]").click()
       driver.find_element_by_xpath("//*[text()='Cucumber - 1 Kg']/following-sibling::div[2]").click()
else:
       driver.find_element_by_xpath("//*[text()='Beetroot - 1 Kg']/following-sibling::div[2]").click()

'''
#4th
'''
driver.get("https://www.makemytrip.com/")
actions = ActionChains(driver)
driver.maximize_window()
driver.find_element_by_xpath("//label[@for='fromCity']/input").click()
time.sleep(2)
driver.find_element_by_xpath("//input[@placeholder='From']").send_keys("goa")
actions.send_keys(Keys.ARROW_DOWN)
actions.perform()
'''
#3rd
'''
driver.get("https://www.spicejet.com/")
driver.maximize_window()
driver.find_element_by_name("ctl00$mainContent$view_date2").click()
if driver.find_element_by_id("Div1").get_attribute("style"):
    print("it is enabled")
else:
    print("it is disabled") 
'''
#2nd:
'''
st=("Hi S will snap in he swan")
for word in st.split():
    if word[0].upper() in {"S","H"}:
         print(word)
'''
#1st:
'''
x=0
while x<5:
    print(x)
    x+=1
'''
