'''''''''''''''''''''''
Python Package Imports
'''''''''''''''''''''''
from time import sleep
from datetime import datetime
import os
import sys
import ctypes
from getpass import getuser
import requests
import openpyxl
from selenium.webdriver.support.ui import Select
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

'''''''''''''''''''''
User Input
'''''''''''''''''''''
pcnam = getuser()
print('\nWelcome '+pcnam)
print('Smallcell Outdoor for Samsung/Airspan Automation Started')
vndr='Abhishek'
print("Current Vendor is : ", vndr)
try:
    vndor = 0
    while not int(vndor) in range(1,3):
        vndor = input("Enter [1] for Samsung or [2] for Airspan : ")
except:
    print("Wrong Values Entered! Kindly Re-run the Program and Enter Proper Values")
    exit()
'''''''''''''''
Path
'''''''''''''''
cwdpath=str(os.getcwd()) #get CWD
cwdpath=cwdpath.replace("\\","/") #Replace Path \''/
'''''''''''''''
'''''''''''''''
if vndor=='1':
    vndr = 'Samsung'
elif vndor=='2':
    vndr = 'Airspan'
print("Vendor Selected is : ", vndr)
'''''''''''''''
SAP ID Check
'''''''''''''''
wb1 = openpyxl.load_workbook(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOImport"+vndr+".xlsx", data_only=True)
wb2 = openpyxl.load_workbook(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOATP11A"+vndr+".xlsx", data_only=True)
if vndor=='1':
    xln="SCOOddNeoSamsung"
else:
    xln="SCOOddAirspan"
wb3 = openpyxl.load_workbook(cwdpath+"/OSC Upload XLs/"+vndr+"/"+xln+".xlsx", data_only=True)
wb4 = openpyxl.load_workbook(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOOdd"+vndr+".xlsx", data_only=True)

if wb4['Sheet1']['A2'].value in {wb1['Sheet1']['U2'].value,wb2['Sheet1']['A2'].value,wb3['Sheet1']['A2'].value}:
    print('SAP IDs in All Excels is Matching... Proceeding...')
    sapid = wb4['Sheet1']['A2'].value
else:
    print('SAP ID in All Excels is MisMatched Kindly Re-run the Program and Enter Proper SAP IDs')
    exit()    

print("\nSAP ID : "+sapid)

'''
print('\nEnter SAP ID: ')
sapid=input()
print('\nSelect which Band to Perform: \n\n1. 2300\n2. 1800\n3. 850 \n\nChoose 1/2/3 : ')
sel=input()
if sel>'3' or sel<'1':
	print("\nPlease Choose Between 1 to 3 only")
	input("\nPress [ENTER] to Exit")
	exit()
'''

'''''''''''''''''''''
Chrome Driver Loads
'''''''''''''''''''''
chrdrv = cwdpath+"/chromedriver.exe"

if Path(chrdrv).is_file():
    print('\nChromeDriver Present Continuing')

else:
    print('\nChromeDriver Absent Cannot Continue')
    print("\nPlease Download ChromeDriver and Place it in your CWD :\n\n"+cwdpath+"/")
    input('\nPress [ENTER] to Exit')
    sys.exit()    

'''
else:
    print('\nChromeDriver Absent Downloading')
    url = "http://www.muhurtnews.com/Abhishek/chromedriver.exe"
    r = requests.get(url, stream = True)
    with open(chrdrv, 'wb') as f:
        for chunk in r.iter_content(chunk_size=1024):
            f.write(chunk)
'''

'''''''''''
Py Drivers
'''''''''''
opts = Options()
opts.add_argument('start-maximized')
opts.add_argument('disable-infobars')
opts.add_argument('--ignore-certificate-errors')
opts.add_argument("--test-type")
driver = webdriver.Chrome(options=opts, executable_path=chrdrv)

"""""""""""""""
Py Functions
"""""""""""""""

'''''''''
Screen
'''''''''
da=datetime.now().strftime('%a_%d-%b-%y')
tim=datetime.now().strftime('%H-%M-%S')
try:
    os.mkdir(r"Scrs")
except:
    pass
try:
    os.mkdir(r"Scrs/"+da)
except:
    pass

def Scr():
    sleep(1)
    driver.save_screenshot("Scrs/"+da+"/"+tim+".png")

'''''''''''''''
Login / Logout
'''''''''''''''
def Login():
    driver.find_element_by_xpath("//*[@class='form-control username']").send_keys(user)
    driver.find_element_by_xpath('//*[@value="Submit"]').click()
    sleep(1)
    driver.find_element_by_xpath("//*[@name='j_password']").send_keys('Pass_123')
    driver.find_element_by_xpath('//*[@id="Login_button"]').click()
    Scr()

def Logout():
    driver.find_element_by_xpath('//*[@id="user-nameco"]').click()
    sleep(1)
    driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[1]/div[2]/div/div/ul/li[3]/a/span').click()

'''''''''''''''''
Code Func()
'''''''''''''''''
def AssBtn():
    driver.find_element_by_xpath('//*[@id="smallcell_out_Atp_11A_task_list_table"]/tbody/tr[2]/td[7]/a').click()
    Scr()

def MKAVen():
    driver.find_element_by_xpath('//*[@id="smallcell_atp11A_task_vendors"]').send_keys('Innoeye')
    Scr()
    driver.find_element_by_xpath('//*[@id="smallcell_atp11A_task_vendorPm"]').send_keys('Amol Gupta')
    Scr()
    driver.find_element_by_xpath('//*[@id="smallcellOut_atp11A_assignTaskToPMButton"]').click()
    Scr()

def AssOKBtn():
    driver.find_element_by_xpath('//*[@id="smallcell_out_Atp_11A_task_list_table"]/tbody/tr[2]/td[7]/a[1]').click()
    Scr()
    driver.find_element_by_xpath('//*[@id="smallcell_Out_ATP11A__accept_task_modal"]/div/div[3]/button[2]').click()
    Scr()

def TskSrchSAP():
    driver.find_element_by_xpath('//*[@id="smallcell_out_atp11a_task_li"]/a').click()
    Scr()
    driver.find_element_by_xpath('//*[@id="smallcell_out_atp_11_A_list_taskSiteSearch"]').send_keys(sapid)
    Scr()
    driver.find_element_by_xpath('//*[@id="small_cell_atp_11_A_TaskSearchTab"]/button').click()
    Scr()

def SCOOddNeo():
    driver.find_element_by_xpath('//*[@id="cellOutdoorSiteId"]/a').click()
    Scr()
    driver.find_element_by_xpath('//*[@id="import_site_odd_for_neo_executive_btn_smc_out_atp11a"]').click()
    Scr()
    driver.find_element_by_xpath("//*[@id='fileUploadInput']").send_keys(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOOddNeo"+vndr+".xlsx") #Upload File
    Scr()
    driver.find_element_by_xpath('//*[@id="smc_out_site_odd_for_neo_executive_triggerUpload"]').click()
    Scr()	

def SCOOddNPE():
    driver.find_element_by_xpath('//*[@id="import_site_btn_title"]').click()
    Scr()
    driver.find_element_by_xpath("//*[@id='fileUploadInput']").send_keys(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOOdd"+vndr+".xlsx") #Upload File
    Scr()
    driver.find_element_by_xpath('//*[@id="smc_out_atp11a_site_triggerUpload"]').click()
    Scr()

def SCOOddAir():
    driver.find_element_by_xpath('//*[@id="import_site_btn_title"]').click()
    Scr()
    driver.find_element_by_xpath("//*[@id='fileUploadInput']").send_keys(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOOdd"+vndr+".xlsx") #Upload File
    Scr()
    driver.find_element_by_xpath('//*[@id="smc_out_atp11a_site_triggerUpload"]').click()
    Scr()

''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

"""""""""""""""
Py Mains
"""""""""""""""
driver.get("http://10.135.26.21:8079/siteforge/jsp/login.jsp") # SF URL Load
sleep(1)

'''
'''''''''''''''''''''''''''''''''''''''''''''''''''
# Smallcell.Admin
'''''''''''''''''''''''''''''''''''''''''''''''''''
#Import Site
user = 'smallcell.admin' #SF Smallcell Outdoor Admin
Login()
sleep(1)
try:
    driver.find_element_by_xpath('//*[@id="small_cell_site_data_upload_modal"]').click() # Click Import Site
    Scr()
    driver.find_element_by_xpath("//*[@id='fileUploadInput']").send_keys(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOImport"+vndr+".xlsx") #Upload File
    Scr()
    driver.find_element_by_xpath('//*[@id="small_cell_site_triggerUpload"]').click() # Click Upload # Commented for Uploading
    Scr()

    #Direct ATP 11A
    driver.find_element_by_xpath('//*[@id="small_cell_atp11A_trigger_data_upload_modal"]').click() # Click ATP Trigger
    Scr()
    driver.find_element_by_xpath("//*[@id='fileUploadInput']").send_keys(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOATP11A"+vndr+".xlsx") #Upload File
    Scr()
    driver.find_element_by_xpath("//input[@id='small_cell_atp11A_task_triggerUpload']").click() # Click Upload # Commented for Uploading
    Scr()
except:
    pass
Logout()

'''''''''''''''''''''''''''''''''''
# Manoj.Kumar
'''''''''''''''''''''''''''''''''''
user = 'manoj.kumar' #SF Smallcell Outdoor Lead
Login()
sleep(1)
driver.find_element_by_xpath('//*[@id="smallcell_out_atp11a_site_li"]/a').click()
Scr()
driver.find_element_by_xpath('//*[@id="smallcellOut_atp11A_site_by_name"]').send_keys(sapid)
Scr()
driver.find_element_by_xpath('//*[@id="container"]/div/div[1]/div/button').click()
Scr()
try:
    driver.find_element_by_xpath('//*[@id="smallcellOut_atp11A_site_list_table"]/tbody/tr[2]/td[5]/a').click()
    Scr()
    driver.find_element_by_xpath('//*[@id="jcEngg"]').send_keys('sachin gupta')
    Scr()
    driver.find_element_by_xpath('//*[@id="smallcell_out_site_assign_to_jc_engineer_modal"]/div/div[3]/a').click()
    Scr()
except:
    pass

TskSrchSAP()
try:
    AssBtn()
    MKAVen()
    driver.find_element_by_xpath('//*[@id="smallcell_out_Atp_11A_task_list_table"]/tbody/tr[3]/td[7]/a').click()
    Scr()
    MKAVen()
except:
    pass
Logout()

'''''''''''''''''''''''''''''''''''
# Amol.Gupta
'''''''''''''''''''''''''''''''''''
user = 'amol.gupta' #SF Smallcell Vendor PM
Login()
sleep(1)
TskSrchSAP()

# Task 1 Equip Inst/Exec ATP11A
try:
    AssBtn()
    driver.find_element_by_xpath('//*[@id="smallcell_Atp11A_task_vendorFE"]').send_keys('Rahul Patidar')
    Scr()
    driver.find_element_by_xpath('//*[@id="smallcell_Atp11A_assignTaskToFeButton"]').click()
    Scr()
except:
    pass

# Task 2 Equip Inst/Exec ATP11A
try:
    driver.find_element_by_xpath('//*[@id="smallcell_out_Atp_11A_task_list_table"]/tbody/tr[3]/td[7]/a').click()
    Scr()
    driver.find_element_by_xpath('//*[@id="smallcell_Atp11A_task_vendorFE"]').send_keys('Rahul Patidar')
    Scr()
    driver.find_element_by_xpath('//*[@id="smallcell_Atp11A_assignTaskToFeButton"]').click()
    Scr()	
except:
    pass
Logout()

'''''''''''''''''''''''''''''''''''
# Rahul.Patidar
'''''''''''''''''''''''''''''''''''
user = 'rahul.patidar' #SF Smallcell FE
Login()
sleep(1)
TskSrchSAP()
sleep(2)
Logout()
ctypes.windll.user32.MessageBoxExW(0, 'Please Perform FE Task Manually as per Requirement and then Press [OK]', 'Khud Se Bharo', 0x1000)

'''''''''''''''''''''''''''''''''''
# Amol.Gupta
'''''''''''''''''''''''''''''''''''
user = 'amol.gupta' #SF Smallcell Vendor PM
Login()
sleep(1)
TskSrchSAP()
try:
    AssOKBtn()
except:
    pass
Logout()

'''''''''''''''''''''''''''''''''''
# Sachin.Gupta
'''''''''''''''''''''''''''''''''''
user = 'sachin.gupta' #SF Smallcell Vendor PM
Login()
sleep(1)
TskSrchSAP()
try:
    AssOKBtn()
except:
    pass
Logout()

'''''''''''''''''''''''''''''''''''
# Manoj.Kumar
'''''''''''''''''''''''''''''''''''
user = 'manoj.kumar' #SF Smallcell Outdoor Lead
Login()
sleep(1)
TskSrchSAP()
try:
    AssOKBtn()	
except:
    pass
Logout()

'''''''''''''''''''''''''''''''''''
# ODD
'''''''''''''''''''''''''''''''''''
if vndor=='1':
    user = 'neo.west' #SF Smallcell Outdoor Lead
    Login()
    SCOOddNeo()
    errmsg = driver.find_element_by_xpath('//*[@id="smc_out_atp11a_site_import_message"]')
    while errmsg.is_displayed():
        print ("\n"+str(errmsg))
        ctypes.windll.user32.MessageBoxExW(0, 'Error while Uploading ODD, Kindly Check and Repair Excel and Press [OK] to Re-Upload', 'Thik Se Bharo', 0x1000)
        driver.find_element_by_xpath('//*[@id="smc_out_atp11a_site_file_upload_modal"]/div/div[1]/button').click()
        SCOOddNeo()
    print ("\nSCO Odd Neo For "+vndr+" Successfully Uploaded")
    Logout()
    
    user = 'mum.npe' #SF Smallcell Outdoor Lead
    Login()
    SCOOddNPE()
    errmsg = driver.find_element_by_xpath('//*[@id="smc_out_atp11a_site_import_message"]')
    while errmsg.is_displayed():
        print ("\n"+str(errmsg))
        ctypes.windll.user32.MessageBoxExW(0, 'Error while Uploading ODD, Kindly Check and Repair Excel and Press [OK] to Re-Upload', 'Thik Se Bharo', 0x1000)
        driver.find_element_by_xpath('//*[@id="smc_out_atp11a_site_file_upload_modal"]/div/div[1]/button').click()
        SCOOddNPE()
    print ("\nSCO Odd NPE For "+vndr+" Successfully Uploaded")
    Logout()

else:
    user = 'airspan' #SF Smallcell Outdoor Lead
    Login()
    SCOOddAir()
    errmsg = driver.find_element_by_xpath('//*[@id="smc_out_atp11a_site_import_message"]')
    while errmsg.is_displayed():
        print ("\n"+str(errmsg))
        ctypes.windll.user32.MessageBoxExW(0, 'Error while Uploading ODD, Kindly Check and Repair Excel and Press [OK] to Re-Upload', 'Thik Se Bharo', 0x1000)
        driver.find_element_by_xpath('//*[@id="smc_out_atp11a_site_file_upload_modal"]/div/div[1]/button').click()
        SCOOddAir()
    print ("\nSCO Odd NPE For "+vndr+" Successfully Uploaded")
    Logout()
'''
'''''''''''''''''''''''''''''''''''
# V.Patil
'''''''''''''''''''''''''''''''''''
user = 'v.patil' #SF Smallcell Outdoor Lead
Login()
sleep(1)
driver.find_element_by_xpath('//*[@id="smallcellOut_atp11A_site_by_name"]').send_keys(sapid)
Scr()
driver.find_element_by_xpath('//*[@id="container"]/div/div[1]/div/button').click()
Scr()
driver.find_element_by_xpath('//*[@id="smallcellOut_atp11A_site_list_table"]/tbody/tr[2]/td[7]/a[1]').click()
Scr()
driver.find_element_by_xpath('//*[@id="odd_accept_button_for_npe"]').click()
Scr()
print ("\nSCO ODD For "+vndr+" Accpeted by NPE Successfully")
Logout()

'''''''''''''''''''''''''''''''''''
# CPNR
'''''''''''''''''''''''''''''''''''
driver.get("http://10.32.131.70:8080/login.jsp") # SF URL Load
sleep(1)

#driver.close()
#driver.quit()
