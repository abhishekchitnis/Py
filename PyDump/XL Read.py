# importing openpyxl module 
import openpyxl 

# Give the location of the file 
path ="C:\\Users\\Abhishek.Chitnis\\Downloads\\Py\@\\demo.xlsx"

# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 

sheet_obj = wb_obj.active

'''
1
'''
# Cell objects also have row, column,  
# and coordinate attributes that provide 
# location information for the cell. 
  
# Note: The first row or  
# column integer is 1, not 0. 
  
# Cell object is created by using  
# sheet object's cell() method. 
cell_obj = sheet_obj.cell(row = 1, column = 1) 
  
# Print value of cell object  
# using the value attribute 
print("1st:",cell_obj.value)

'''
2
'''
# print the total number of rows 
print("\n2nd:",sheet_obj.max_row) 

'''
3
'''
# print the total number of columns 
print("\n3rd:",sheet_obj.max_column)

'''
4
'''  
# Loop will print all columns name 
max_col = sheet_obj.max_column
print("\n4th:")
for i in range(1, max_col + 1): 
    cell_obj = sheet_obj.cell(row = 1, column = i) 
    print(cell_obj.value)  

'''
5
'''
m_row = sheet_obj.max_row 
print("\n5th:")
# Loop will print all values 
# of first column  
for i in range(1, m_row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 1) 
    print(cell_obj.value)

'''
6
'''
max_col = sheet_obj.max_column 
print("\n6th:")  
# Will print a particular row value 
for i in range(1, max_col + 1): 
    cell_obj = sheet_obj.cell(row = 2, column = i) 
    print(cell_obj.value, end =" | ")


print("")  
'''
End
'''