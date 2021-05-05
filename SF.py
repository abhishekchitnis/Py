'''''''''''''''''''''''
Python Package Imports
'''''''''''''''''''''''
from time import sleep
from datetime import datetime
import os
import sys
import ctypes
from getpass import getuser
import requests
import openpyxl
from selenium.webdriver.support.ui import Select
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

'''''''''''''''''''''
# User Input
'''''''''''''''''''''
pcnam = getuser()
aousrnm = 'Abhishek'
print('Welcome '+pcnam+' to UMS Automation Script')
try:
        aousr = input("Enter [1] for AD User or [2] for OID User : ")
        if aousr=='1':
                aousrnm = 'AD User'
                print("User Selected for Creation is : ", aousrnm)
        elif aousr=='2':
                aousrnm = 'OID User'
                print("User Selected for Creation is : ", aousrnm)
        else:
                print("Wrong Values Entered! Kindly Re-run the Program and Enter Proper Values")
                exit()
except:
        exit()
'''''''''''''''
# Path
'''''''''''''''
cwdpath=str(os.getcwd()) #get CWD
cwdpath=cwdpath.replace("\\","/") #Replace Path \''/

'''''''''''''''
# SAP ID Check
'''''''''''''''
'''
wb1 = openpyxl.load_workbook(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOImport"+vndr+".xlsx", data_only=True)
wb2 = openpyxl.load_workbook(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOATP11A"+vndr+".xlsx", data_only=True)
if vndor=='1':
    xln="SCOOddNeoSamsung"
else:
    xln="SCOOddAirspan"
wb3 = openpyxl.load_workbook(cwdpath+"/OSC Upload XLs/"+vndr+"/"+xln+".xlsx", data_only=True)
wb4 = openpyxl.load_workbook(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOOdd"+vndr+".xlsx", data_only=True)

if wb4['Sheet1']['A2'].value in {wb1['Sheet1']['U2'].value,wb2['Sheet1']['A2'].value,wb3['Sheet1']['A2'].value}:
    print('SAP IDs in All Excels is Matching... Proceeding...')
    sapid = wb4['Sheet1']['A2'].value
else:
    print('SAP ID in All Excels is MisMatched Kindly Re-run the Program and Enter Proper SAP IDs')
    exit()    

print("\nSAP ID : "+sapid)


print('\nEnter SAP ID: ')
sapid=input()
print('\nSelect which Band to Perform: \n\n1. 2300\n2. 1800\n3. 850 \n\nChoose 1/2/3 : ')
sel=input()
if sel>'3' or sel<'1':
	print("\nPlease Choose Between 1 to 3 only")
	input("\nPress [ENTER] to Exit")
	exit()
'''

'''''''''''''''''''''
# Chrome Driver Loads
'''''''''''''''''''''
chrdrv = cwdpath+"/chromedriver.exe"

if Path(chrdrv).is_file():
        print('\nChromeDriver Present Continuing')

else:
        print('\nChromeDriver Absent Cannot Continue')
        print("\nPlease Download ChromeDriver and Place it in your CWD :\n\n"+cwdpath+"/")
        input('\nPress [ENTER] to Exit')
        exit()    

'''
else:
    print('\nChromeDriver Absent Downloading')
    url = "http://www.muhurtnews.com/Abhishek/chromedriver.exe"
    r = requests.get(url, stream = True)
    with open(chrdrv, 'wb') as f:
        for chunk in r.iter_content(chunk_size=1024):
            f.write(chunk)
'''

'''''''''''
Py Drivers
'''''''''''
opts = Options()
opts.add_argument('start-maximized')
opts.add_argument('disable-infobars')
opts.add_argument('--ignore-certificate-errors')
opts.add_argument("--test-type")
driver = webdriver.Chrome(options=opts, executable_path=chrdrv)

"""""""""""""""
Py Functions
"""""""""""""""

'''''''''
Screen
'''''''''
da=datetime.now().strftime('%a_%d-%b-%y')
tim=datetime.now().strftime('%H-%M-%S')
try:
        os.mkdir(r"Scrs")
except:
        pass
try:
        os.mkdir(r"Scrs/"+da)
except:
        pass

def Scr():
        sleep(1)
        driver.save_screenshot("Scrs/"+da+"/"+tim+".png")

'''''''''''''''
Login / Logout
'''''''''''''''
def Login():
        user='Abhishek.Chitnis'
        driver.find_element_by_xpath('//*[@id="content-inner"]/form/input[3]').send_keys(user)
        pwd='0%(#arGED'
        driver.find_element_by_xpath('//*[@id="content-inner"]/form/input[4]').send_keys(pwd)
        driver.find_element_by_xpath('//*[@id="content-inner"]/form/input[5]').click()
        sleep(1)
        Scr()

def Logout():
        driver.find_element_by_xpath('//*[@id="user-nameco"]').click()
        sleep(1)
        driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[1]/div[2]/div/div/ul/li[3]/a/span').click()

'''''''''''''''''
Code Func()
'''''''''''''''''
def AssBtn():
    driver.find_element_by_xpath('//*[@id="smallcell_out_Atp_11A_task_list_table"]/tbody/tr[2]/td[7]/a').click()
    Scr()

''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

"""""""""""""""
Py Mains
"""""""""""""""
driver.get("http://10.135.26.21:8081/ums") # UMS URL Load
sleep(.5)

'''''''''''''''''''''''''''''''''''''''''''''''''''
# UMS Create User
'''''''''''''''''''''''''''''''''''''''''''''''''''
Login()
driver.find_element_by_xpath('//*[@id="userFunctionCard"]/div[1]/div/div[2]/a').click()
try:
        if aousr==1:
                driver.find_element_by_xpath('//*[@id="createADUserID"]').send_keys('Abhishek.Chitnis')
                driver.find_element_by_xpath('//*[@id="userTypeCard"]/div[1]/div/div[2]/div[2]/div/a').click()
                
        else:
                driver.find_element_by_xpath('//*[@id="userTypeCard"]/div[2]/div/div[2]/div/a').click()
                
        driver.get("http://10.135.26.21:8079/siteforge/") # UMS URL Load
        #sleep(.5)
        driver.find_element_by_xpath('//*[@id="domainSelect"]').send_keys('UBR')
        driver.find_element_by_xpath('//*[@id="loginSubmitBtn"]').click()
        ctypes.windll.user32.MessageBoxExW(0, 'Please Perform FE Task Manually as per Requirement and then Press [OK]', 'Khud Se Bharo', 0x1000)
                
        
except Exception as e: print(e)

'''
try:
    driver.find_element_by_xpath('//*[@id="small_cell_site_data_upload_modal"]').click() # Click Import Site
    Scr()
    driver.find_element_by_xpath("//*[@id='fileUploadInput']").send_keys(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOImport"+vndr+".xlsx") #Upload File
    Scr()
    driver.find_element_by_xpath('//*[@id="small_cell_site_triggerUpload"]').click() # Click Upload # Commented for Uploading
    Scr()

    #Direct ATP 11A
    driver.find_element_by_xpath('//*[@id="small_cell_atp11A_trigger_data_upload_modal"]').click() # Click ATP Trigger
    Scr()
    driver.find_element_by_xpath("//*[@id='fileUploadInput']").send_keys(cwdpath+"/OSC Upload XLs/"+vndr+"/SCOATP11A"+vndr+".xlsx") #Upload File
    Scr()
    driver.find_element_by_xpath("//input[@id='small_cell_atp11A_task_triggerUpload']").click() # Click Upload # Commented for Uploading
    Scr()
except:
    pass
Logout()
'''

#driver.close()
#driver.quit()
